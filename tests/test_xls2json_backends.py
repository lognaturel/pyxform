"""Test xls2json_backends module functionality."""

import datetime
import os
from pathlib import Path
from unittest import expectedFailure

import openpyxl
from pyxform.builder import create_survey_element_from_dict
from pyxform.xls2json import workbook_to_json
from pyxform.xls2json_backends import (
    convert_file_to_csv_string,
    csv_to_dict,
    get_xlsform,
    md_table_to_workbook,
    md_to_dict,
    xlsx_to_dict,
    xlsx_value_to_str,
)

from tests import utils
from tests.fixtures import bug_example_forms
from tests.pyxform_test_case import PyxformTestCase
from tests.xpath_helpers.choices import xpc
from tests.xpath_helpers.entities import xpe
from tests.xpath_helpers.questions import xpq
from tests.xpath_helpers.settings import xps


class TestXLS2JSONBackends(PyxformTestCase):
    """Test xls2json_backends module."""

    maxDiff = None

    def test_xlsx_value_to_str(self):
        value = 32.0
        csv_data = xlsx_value_to_str(value)
        expected_output = "32"
        self.assertEqual(csv_data, expected_output)

        # Test that the decimal value is not changed during conversion.
        value = 46.9
        csv_data = xlsx_value_to_str(value)
        expected_output = "46.9"
        self.assertEqual(csv_data, expected_output)

    def test_defusedxml_enabled(self):
        self.assertTrue(openpyxl.DEFUSEDXML)

    def test_case_insensitivity(self):
        """Should find all input types are case-insensitive for sheet names and headers."""
        # Exhaustive matches for XLSForm content to check all sheets / headers work.
        xml__xpath_match = [
            # survey
            xpq.model_instance_item("q1"),
            xpq.model_itext_label("q1", "EN", "Are you good?"),
            xpq.body_control("q1", "select1"),
            xpq.model_instance_item("q2"),
            xpq.model_itext_label("q2", "EN", "Where are you?"),
            xpq.body_control("q2", "input"),
            xpq.model_instance_item("q3"),
            xpq.model_itext_label("q3", "EN", "Where exactly?"),
            xpq.body_control("q3", "upload"),
            # choices
            xpc.model_instance_choices_itext("c1", ("n1-c", "n2-c")),
            xpc.model_itext_choice_text_label_by_pos("EN", "c1", ("l1-c", "l2-c")),
            # settings
            xps.form_title("Yes or no"),
            xps.form_id("YesNo"),
            xps.language_is_default("EN"),
            # external choices
            """
            /h:html/h:body/x:input[
            @ref='/test_name/q2'
            and @query="instance('c1')/root/item[YES_NO= /test_name/q1 ]"
            ]
            """,
            # entities
            xpe.model_instance_meta("e1", create=True, label=True),
            xpe.model_bind_meta_label("l1"),
            # osm
            xpq.body_upload_tags("q3", (("n1-o", "l1-o"), ("n2-o", "l2-o"))),
        ]
        file_types = [".xlsx", ".csv", ".md"]
        for file_type in file_types:
            with self.subTest(msg=file_type):
                data = get_xlsform(
                    xlsform=utils.path_to_text_fixture(f"case_insensitivity{file_type}"),
                )
                # All sheets recognised.
                for attr in data.__slots__:
                    self.assertIsNotNone(getattr(data, attr))
                # Expected original sheet_names - needed for spellchecks.
                self.assertEqual(
                    [
                        "SURVEY",
                        "CHOICES",
                        "SETTINGS",
                        "EXTERNAL_CHOICES",
                        "ENTITIES",
                        "OSM",
                    ],
                    data.sheet_names,
                )
                # Headers stripped, but not split or lower-cased yet, since this requires
                # more complex logic that is part of workbook_to_json.
                self.assertEqual(
                    ["TYPE", "NAME", "LABEL::EN", "CHOICE_FILTER"],
                    list(data.survey_header[0].keys()),
                )
                self.assertEqual(
                    ["LIST_NAME", "NAME", "LABEL::EN"],
                    list(data.choices_header[0].keys()),
                )
                self.assertEqual(
                    ["FORM_TITLE", "FORM_ID", "DEFAULT_LANGUAGE"],
                    list(data.settings_header[0].keys()),
                )
                self.assertEqual(
                    ["LIST_NAME", "NAME", "LABEL", "YES_NO"],
                    list(data.external_choices_header[0].keys()),
                )
                self.assertEqual(
                    ["DATASET", "LABEL"], list(data.entities_header[0].keys())
                )
                self.assertEqual(
                    ["LIST_NAME", "NAME", "LABEL"], list(data.osm_header[0].keys())
                )
                # All columns recognised.
                pyxform = workbook_to_json(workbook_dict=data, form_name="test_name")
                self.assertPyxformXform(
                    survey=create_survey_element_from_dict(pyxform),
                    xml__xpath_match=xml__xpath_match,
                )

    def test_equivalency(self):
        """Should get the same data from equivalent files using each file type reader."""
        equivalent_fixtures = [
            "group",
            "loop",
            "text_and_integer",
            "yes_or_no_question",
        ]
        for fixture in equivalent_fixtures:
            xlsx_inp = xlsx_to_dict(utils.path_to_text_fixture(f"{fixture}.xlsx"))
            csv_inp = csv_to_dict(utils.path_to_text_fixture(f"{fixture}.csv"))
            md_inp = md_to_dict(utils.path_to_text_fixture(f"{fixture}.md"))

            self.assertEqual(xlsx_inp, csv_inp)
            self.assertEqual(xlsx_inp, md_inp)

    def test_form_with_many_empty_cells(self):
        """Should quickly produce expected data, and find large input sheet dimensions."""
        # Test fixture produced by adding data at cells IV1 and A19999.
        xlsx_path = os.path.join(bug_example_forms.PATH, "extra_columns.xlsx")
        before = datetime.datetime.now(datetime.UTC)
        xlsx_data = xlsx_to_dict(xlsx_path)
        after = datetime.datetime.now(datetime.UTC)
        self.assertLess((after - before).total_seconds(), 5)
        wb = openpyxl.open(filename=xlsx_path, read_only=True, data_only=True)

        survey_headers = [
            "type",
            "name",
            "label",
        ]
        self.assertEqual(survey_headers, list(xlsx_data["survey_header"][0].keys()))
        self.assertEqual(3, len(xlsx_data["survey"]))
        self.assertEqual("b", xlsx_data["survey"][2]["name"])
        survey = wb["survey"]
        self.assertEqual(19999, survey.max_row)
        self.assertEqual(256, survey.max_column)

        wb.close()

    def test_xlsx_with_many_empty_rows(self):
        """Should quickly produce expected data, and find large input sheet dimensions."""
        # Test fixture produced (presumably) by a LibreOffice serialisation bug.
        xlsx_path = os.path.join(bug_example_forms.PATH, "UCL_Biomass_Plot_Form.xlsx")
        before = datetime.datetime.now(datetime.UTC)
        xlsx_data = xlsx_to_dict(xlsx_path)
        after = datetime.datetime.now(datetime.UTC)
        self.assertLess((after - before).total_seconds(), 5)
        wb = openpyxl.open(filename=xlsx_path, read_only=True, data_only=True)

        survey_headers = [
            "type",
            "name",
            "label::Swahili (sw)",
            "label::English (en)",
            "hint::Swahili (sw)",
            "hint::English (en)",
            "required",
            "relevant",
            "constraint",
            "constraint_message::Swahili (sw)",
            "constraint_message::English (en)",
            "choice_filter",
            "appearance",
            "calculation",
            "repeat_count",
            "parameters",
        ]
        # Expected headers, rows, and last row contains expected data.
        self.assertEqual(survey_headers, list(xlsx_data["survey_header"][0].keys()))
        self.assertEqual(90, len(xlsx_data["survey"]))
        self.assertEqual("deviceid", xlsx_data["survey"][89]["type"])
        survey = wb["survey"]
        self.assertTupleEqual((1048576, 1024), (survey.max_row, survey.max_column))

        choices_headers = [
            "list_name",
            "name",
            "label::Swahili (sw)",
            "label::English (en)",
        ]
        self.assertEqual(choices_headers, list(xlsx_data["choices_header"][0].keys()))
        self.assertEqual(27, len(xlsx_data["choices"]))
        self.assertEqual("Mwingine", xlsx_data["choices"][26]["label::Swahili (sw)"])
        choices = wb["choices"]
        self.assertTupleEqual((1048576, 4), (choices.max_row, choices.max_column))

        settings_headers = [
            "default_language",
            "version",
        ]
        self.assertEqual(settings_headers, list(xlsx_data["settings_header"][0].keys()))
        self.assertEqual(1, len(xlsx_data["settings"]))
        self.assertEqual("Swahili (sw)", xlsx_data["settings"][0]["default_language"])
        settings = wb["settings"]
        self.assertTupleEqual((2, 2), (settings.max_row, settings.max_column))

        wb.close()

    def test_convert_file_to_csv_string(self):
        """Should find that the same data read from xlsx or csv results in the same csv."""
        md = """
        | survey |
        | | type | name | label |
        | | text | q1   | Q1    |
        """
        csv = "survey\n,type,name,label\n,text,q1,Q1"
        wb = md_table_to_workbook(md)
        with utils.get_temp_dir() as tmp:
            form_path = Path(tmp) / "test_name.xlsx"
            wb.save(form_path)
            wb.close()
            converted_form = convert_file_to_csv_string(str(form_path))
            csv_path = Path(tmp) / "test_name.csv"
            csv_path.write_text(csv)
            converted_csv = convert_file_to_csv_string(str(csv_path))
        self.assertEqual(converted_csv, converted_form)

    @expectedFailure
    def test_convert_file_to_csv_string__column_order_bug(self):
        """Should find that the header order is preserved when writing to csv."""
        md = """
        | survey |
        | | type | name | label | hint | constraint | relevant |
        | | text | q1   | Q1    | Hmm  |            | true()   |
        | | text | q2   | Q2    |      | . != '0'   |          |
        """
        csv = (
            "survey"
            "\n,type,name,label,hint,constraint,relevant"
            "\n,text,q1,Q1,Hmm,,true()"
            "\n,text,q2,Q2,,\". != '0'\","
        )
        wb = md_table_to_workbook(md)
        with utils.get_temp_dir() as tmp:
            form_path = Path(tmp) / "test_name.xlsx"
            wb.save(form_path)
            wb.close()
            converted_form = convert_file_to_csv_string(str(form_path))
            csv_path = Path(tmp) / "test_name.csv"
            csv_path.write_text(csv)
            converted_csv = convert_file_to_csv_string(str(csv_path))
        self.assertEqual(converted_csv, converted_form)

        # Bug seems to occur because convert_file_to_csv collects the headers from the
        # rows rather than using the "survey_header" (or similar). In this case the first
        # row has no constraint, but the second row does, so constraint is put last.
        expected_header = ",type,name,label,hint,constraint,relevant"
        bugged_header = ",type,name,label,hint,relevant,constraint"
        # Expected header is there but only in the the "survey_header" dict.
        self.assertEqual(1, converted_csv.count(expected_header))
        self.assertEqual(1, converted_form.count(expected_header))
        # Bugged header appears in the "survey" dict.
        self.assertEqual(1, converted_csv.count(bugged_header))
        self.assertEqual(1, converted_form.count(bugged_header))
        # Fail the test since the "survey" data should retain column order.
        self.assertEqual(2, converted_csv.count(expected_header))
        self.assertEqual(2, converted_form.count(expected_header))
